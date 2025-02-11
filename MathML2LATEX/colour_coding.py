

# Importing the libraries

from pinecone import Pinecone
from sentence_transformers import SentenceTransformer
import re
import ast
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

#%%

#Connecting to the Pinecone Database

pc = Pinecone(api_key="7c4917bc-322f-4d75-80af-980c6628f46f")
index = pc.Index("infinity-learn-6-10")

#Selecting the embedding model
model = SentenceTransformer('paraphrase-MiniLM-L6-v2')

#%%

# Function for RAG finding
def RAG(question):
    
    #creating the embeddings for the asked question/doubt
    embedding1 = model.encode(question)
    
    #finding the top 5 relevant ids on the basis of vector score
    answer =index.query(
          namespace="",
          vector=embedding1.tolist(),
          top_k=5,
    # include_values=True,
         include_metadata=True) 
    
    #iterating through the metadata to get the question, answer options, correct answer, solution and score
    question_list=[]
    for i in range(0,5):
        try:

            if "solution " in answer['matches'][i]['metadata'].keys():
                try:
                    question_list.append({
                        "ID":answer['matches'][i]['id'],
                        "Score":answer['matches'][i]['score'],
                        #"Difficulty Level":answer['matches'][i]['metadata']["Difficulty Level"],
                        #"question_type":answer['matches'][i]['metadata']["question_type"],
                        #"questiontype":answer['matches'][i]['metadata']["questiontype"],
                        "Question":answer['matches'][i]['metadata']["question_text"],
                        "Answer_Options":answer['matches'][i]['metadata']["answer_options"],
                        #"Solution_Data":ast.literal_eval(answer['matches'][i]['metadata']['solution '])[0]["data"]
                        
                        
                            
                        })
                except:
                    pattern1 = r'<p.*/p>'
                    question_list.append({
                        "ID":answer['matches'][i]['id'],
                        "Score":answer['matches'][i]['score'],
                        #"Difficulty Level":answer['matches'][i]['metadata']["Difficulty Level"],
                        #"question_type":answer['matches'][i]['metadata']["question_type"],
                        #"questiontype":answer['matches'][i]['metadata']["questiontype"],
                        "Question":answer['matches'][i]['metadata']["question_text"],
                        "Answer_Options":answer['matches'][i]['metadata']["answer_options"],
                        #"solution_data":re.search(pattern1, answer['matches'][i]['metadata']['solution '])[0]
                        
                        
                        })
        #if any question is not getting retrieved due to some issues, get the error message            
        except:
            print("invalid question RAG")
        
    return question_list

#%%

def extract_questions(question,count):

    output=RAG(question)
    #print (output)
    
    #list to store individual questions
    questions=[]
    scores=[]
    
    # Extracting questions and storing them in separate variables
    for i in range (5):
        if output[i] and (output[i]["Score"]>0.65):
            questions.append(output[i]["Question"])
            
        else:
                questions.append(output[i]["Score"])
    
        scores.append(output[i]["Score"])
        print(i)
    
    while len(questions) <5:
        questions.append("")
    
    
    #print (questions)
    print(count)
    print(scores)
    


    return questions,scores
       
        

#%%
def color_code(input_file, output_file):
    # Import excel
    df = pd.read_excel(input_file, engine='openpyxl')
    
    # DataFrame for scores
    scores_df = pd.DataFrame(columns=["Score1", "Score2", "Score3", "Score4", "Score5"])
    
    count = 0
    for index, row in df.iterrows():
        question_text = row["Question"]
        _, scores = extract_questions(question_text, count)
        
        for i in range(5):
            scores_df.at[index, f"Score{i+1}"] = scores[i]
        count += 1
    
    print(scores_df)
    print(count)
    
    # Load the workbook to apply conditional formatting
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # Define the colors
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    green_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Apply the conditional formatting based on the scores
    for index, row in scores_df.iterrows():
        for col, score in enumerate(row, start=6):  # Start from column B (2)
            cell = ws.cell(row=index+2, column=col)  # +2 because Excel is 1-indexed and we have a header row
            try:
                score = float(score)
                if score < 40:
                    cell.fill = red_fill
                elif 40 <= score < 65:
                    cell.fill = orange_fill
                elif 65 <= score < 75:
                    cell.fill = green_fill
                elif score >= 75:
                    cell.fill = yellow_fill
            except (ValueError, TypeError):
                # Ignore cells that do not contain a score
                continue
    
    # Save the workbook with conditional formatting applied
    wb.save(output_file)
    print(f"Color coding applied and saved to {output_file}")
    
    return scores_df


inputfile ="test1.xlsx"
outputfile = "test2.xlsx"
color_code(inputfile,outputfile)



































